#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
A multi-domain price scraper for e-commerce sites.

This script can be run in two ways:
1. Single URL Mode: python3 priceScrap.py 'http://example.com/product/123'
2. Batch Mode: python3 priceScrap.py 'targetURL.json'

It scrapes product data (SKU, title, price), saves an HTML snapshot,
and logs the price history in two separate Excel files: one for the
individual item's history and one for the domain's overall summary.
"""

import sys
import os
import json
import re
import datetime
import requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- Helper Functions ---

def get_domain_name(url):
    """Extracts the simple domain name (e.g., 'nova.ge') from a URL."""
    try:
        parsed_url = urlparse(url)
        domain_parts = parsed_url.netloc.split('.')
        if len(domain_parts) >= 2:
            # Join the last two parts, e.g., 'www.domino.com.ge' -> 'domino.com.ge'
            return '.'.join(domain_parts[-(len(domain_parts)-1) if 'www' in domain_parts[0] else -2:])
        return parsed_url.netloc
    except Exception as e:
        print(f"Error parsing domain from URL: {url} - {e}")
        return None

def clean_price(price_str):
    """Converts a price string (e.g., '164,00 ₾') to a float (e.g., 164.00)."""
    if price_str is None:
        return None
    try:
        # Remove currency symbols, whitespace, and replace comma with a dot
        cleaned_str = re.sub(r'[^\d,\.]', '', price_str).replace(',', '.')
        return float(cleaned_str)
    except (ValueError, TypeError):
        return None

def get_modification_time():
    """Returns the current date and time as a string."""
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# --- Parser Functions ---

def parse_nova_ge(soup):
    """Scrapes product data from a nova.ge BeautifulSoup object."""
    data = {'sku': None, 'item_id': None, 'new_price': None, 'old_price': None, 'title': None, 'error': None}
    
    try:
        # 1. Find SKU and Item ID
        sku_parent_element = soup.find(class_="sku")
        if not sku_parent_element:
            data['error'] = "Could not find SKU parent element with class='sku'."
            return data
            
        sku_span = sku_parent_element.find("span", id=re.compile(r"^sku-"))
        if not sku_span:
            data['error'] = "Could not find SKU span with id starting with 'sku-'."
            return data
            
        data['sku'] = sku_span.get_text(strip=True)
        data['item_id'] = sku_span['id'].replace('sku-', '')
        
        # 2. Find Prices using the Item ID
        old_price_tag = soup.find(class_=f"product__oldprice old-price-value-{data['item_id']}")
        new_price_tag = soup.find(class_=f"product__newprice price-value-{data['item_id']}")
        
        if new_price_tag:
            data['new_price'] = clean_price(new_price_tag.get_text(strip=True))
            if old_price_tag:
                data['old_price'] = clean_price(old_price_tag.get_text(strip=True))
            else:
                # If there's a new price but no old price, it's not a sale
                data['old_price'] = None
        else:
            # If no new_price_tag, the main price is the old_price_tag (or a single non-sale price)
            # This logic might need adjustment based on non-sale items
            single_price_tag = soup.find(class_=re.compile(r"price-value-"))
            if single_price_tag:
                 data['new_price'] = clean_price(single_price_tag.get_text(strip=True))
            data['old_price'] = None

        # 3. Find Title
        title_tag = soup.find(class_="product__details--title")
        if title_tag and title_tag.h1:
            data['title'] = title_tag.h1.get_text(strip=True)
        else:
            data['error'] = "Could not find product title."
            return data
            
        return data

    except Exception as e:
        data['error'] = f"An unexpected error occurred during nova.ge parsing: {e}"
        return data

def parse_domino_com_ge(soup):
    """Scrapes product data from a domino.com.ge BeautifulSoup object."""
    data = {'sku': None, 'item_id': None, 'new_price': None, 'old_price': None, 'title': None, 'error': None}
    
    try:
        # 1. Find SKU and Item ID
        sku_span = soup.find("span", id=re.compile(r"^product_code_"))
        if not sku_span:
            data['error'] = "Could not find SKU span with id starting with 'product_code_'."
            return data

        # Get SKU by finding the first text node, ignoring comments
        data['sku'] = sku_span.find(string=True, recursive=False).strip()
        
        # Get Item ID from the span's 'id' attribute
        data['item_id'] = sku_span['id'].replace('product_code_', '')

        # 2. Find Prices using the Item ID
        new_price_tag = soup.find("span", id=f"sec_discounted_price_{data['item_id']}")
        old_price_tag = soup.find("span", id=f"sec_list_price_{data['item_id']}")

        if new_price_tag:
            # This means it's a sale item
            data['new_price'] = clean_price(new_price_tag.get_text(strip=True))
            if old_price_tag:
                data['old_price'] = clean_price(old_price_tag.get_text(strip=True))
        elif old_price_tag:
            # This means it's not on sale, and 'old_price_tag' is the main price
            data['new_price'] = clean_price(old_price_tag.get_text(strip=True))
            data['old_price'] = None
        else:
            # Fallback if only one price (not 'list' or 'discounted') is present
            price_tag = soup.find("span", id=re.compile(r"price_"))
            if price_tag:
                data['new_price'] = clean_price(price_tag.get_text(strip=True))
            data['old_price'] = None

        # 3. Find Title
        title_container = soup.find(class_="ut2-pb__title")
        if title_container and title_container.h1 and title_container.h1.bdi:
            data['title'] = title_container.h1.bdi.get_text(strip=True)
        elif title_container and title_container.h1: # Fallback
            data['title'] = title_container.h1.get_text(strip=True)
        else:
            data['error'] = "Could not find product title."
            return data
            
        return data

    except Exception as e:
        data['error'] = f"An unexpected error occurred during domino.com.ge parsing: {e}"
        return data

# --- Excel Functions ---

def style_excel_headers(ws):
    """Applies bold font and light grey fill to the first row of a worksheet."""
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

def adjust_excel_columns(ws):
    """Adjusts column widths for better readability."""
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 50: # Cap width
            adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width

def update_item_excel(item_dir, sku, item_id, title, new_price, old_price, url, mod_time):
    """
    Updates or creates the item-specific Excel log.
    Logs every scrape attempt.
    """
    excel_file = os.path.join(item_dir, f"sku-{item_id}.xlsx")
    headers = ["SKU", "newPrice", "oldPrice", "title", "URL", "scrapeTime"]
    
    # Define the green fill for sale prices
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    try:
        if not os.path.exists(excel_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PriceHistory"
            ws.append(headers)
            style_excel_headers(ws)
        else:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        
        # Add new data
        row_data = [sku, new_price, old_price, title, url, mod_time]
        ws.append(row_data)
        
        # Style the new row
        new_row_idx = ws.max_row
        if old_price and new_price:
            ws[f'B{new_row_idx}'].fill = green_fill # Style 'newPrice'
        
        adjust_excel_columns(ws)
        wb.save(excel_file)
        
    except Exception as e:
        print(f"Error updating item Excel '{excel_file}': {e}")

def update_domain_excel(domain_dir, domain_name, sku, item_id, title, new_price, old_price, url, mod_time):
    """
    Updates or creates the master domain summary Excel file.
    Keeps only the latest data, min/max prices.
    """
    excel_file = os.path.join(domain_dir, f"{domain_name}-summary.xlsx")
    headers = [
        "SKU", "item_ID", "newPrice", "oldPrice", "title", 
        "minPrice", "maxPrice", "URL", "lastModified"
    ]
    
    # Define the green fill for sale prices
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    try:
        if not os.path.exists(excel_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ProductSummary"
            ws.append(headers)
            style_excel_headers(ws)
        else:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
        # Find existing row for this item_ID
        target_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=2).value == item_id: # Column B is item_ID
                target_row = row
                break
        
        current_min_price = None
        current_max_price = None
        
        if target_row:
            # Get existing min/max prices
            current_min_price = ws.cell(row=target_row, column=6).value
            current_max_price = ws.cell(row=target_row, column=7).value
        
        # Determine new min/max
        if new_price is not None:
            new_min = min(current_min_price, new_price) if current_min_price is not None else new_price
            new_max = max(current_max_price, new_price) if current_max_price is not None else new_price
        else:
            new_min = current_min_price
            new_max = current_max_price

        # Prepare new row data
        row_data = [
            sku, item_id, new_price, old_price, title,
            new_min, new_max, url, mod_time
        ]

        if target_row:
            # Update existing row
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=target_row, column=col_idx, value=value)
        else:
            # Add new row
            ws.append(row_data)
            target_row = ws.max_row
            
        # Style the row
        if old_price and new_price:
            ws[f'C{target_row}'].fill = green_fill # Style 'newPrice'
        else:
            ws[f'C{target_row}'].fill = PatternFill(fill_type=None) # Clear fill if not on sale

        adjust_excel_columns(ws)
        wb.save(excel_file)

    except Exception as e:
        print(f"Error updating domain Excel '{excel_file}': {e}")


# --- Main Scraper Function ---

def process_url(url):
    """
    Main function to process a single URL.
    - Fetches HTML
    - Calls the correct domain parser
    - Saves HTML
    - Updates Excel files
    """
    print(f"\n--- Processing: {url} ---")
    
    try:
        # 1. Get Domain
        domain_name = get_domain_name(url)
        if not domain_name:
            print(f"FAILED: Could not determine domain for {url}")
            return False

        # 2. Fetch HTML
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(url, headers=headers)
        response.raise_for_status() # Raise HTTPError for bad responses (4xx or 5xx)
        html_content = response.text
        soup = BeautifulSoup(html_content, 'html.parser')

        # 3. Call Correct Parser based on domain
        scraped_data = None
        if domain_name == 'nova.ge':
            scraped_data = parse_nova_ge(soup)
        elif domain_name == 'domino.com.ge':
            scraped_data = parse_domino_com_ge(soup)
        else:
            print(f"FAILED: No parser available for domain: {domain_name}")
            return False
            
        # 4. Check for parsing errors
        if scraped_data.get('error'):
            print(f"A parsing error occurred: {scraped_data['error']}")
            raise Exception(scraped_data['error']) # Raise to trigger failure message
            
        sku = scraped_data['sku']
        item_id = scraped_data['item_id']
        new_price = scraped_data['new_price']
        old_price = scraped_data['old_price']
        title = scraped_data['title']
        
        if not all([sku, item_id, title]):
             raise Exception("Missing essential data (SKU, Item ID, or Title) after parsing.")

        print(f"  > Found SKU: {sku}")
        print(f"  > Found Item ID: {item_id}")
        print(f"  > Found Title: {title}")
        print(f"  > Found Price: {new_price} (Old: {old_price})")

        # 5. Create Directories
        base_dir = os.path.dirname(os.path.abspath(__file__))
        domain_dir = os.path.join(base_dir, domain_name)
        item_dir = os.path.join(domain_dir, item_id)
        os.makedirs(item_dir, exist_ok=True)
        
        mod_time = get_modification_time()
        mod_time_file_safe = mod_time.replace(':', '-').replace(' ', '_')

        # 6. Save HTML
        html_filename = f"sku-{item_id}_{mod_time_file_safe}.html"
        html_filepath = os.path.join(item_dir, html_filename)
        with open(html_filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)

        # 7. Update Excel Files
        # Update the item-specific log
        update_item_excel(item_dir, sku, item_id, title, new_price, old_price, url, mod_time)
        
        # Update the master domain summary
        update_domain_excel(domain_dir, domain_name, sku, item_id, title, new_price, old_price, url, mod_time)

        print(f"SUCCESS: Successfully processed and logged data for Item ID {item_id}.")
        return True

    except requests.exceptions.HTTPError as e:
        print(f"HTTP Error: {e.response.status_code} for URL: {url}")
    except requests.exceptions.RequestException as e:
        print(f"Request Error: {e} for URL: {url}")
    except Exception as e:
        print("\n" + "="*50)
        print("--- SCRIPT FAILED ---")
        print(f"An error occurred: {e}")
        print("This usually means the website's HTML structure has changed,")
        print("and the script couldn't find the required elements (like 'sku', 'price', etc.).")
        print("="*50 + "\n")
    
    return False

# --- Main Execution ---

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python3 priceScrap.py '<URL>' OR 'targetURL.json'")
        sys.exit(1)

    argument = sys.argv[1]

    if argument.endswith('.json'):
        # Batch Mode
        print(f"Batch mode activated. Reading from '{argument}'...")
        try:
            with open(argument, 'r') as f:
                urls = json.load(f)
            
            if not isinstance(urls, list):
                print(f"Error: JSON file '{argument}' must contain a list of objects.")
                sys.exit(1)

            total = len(urls)
            success_count = 0
            
            for i, item in enumerate(urls, 1):
                if not isinstance(item, dict) or 'url' not in item:
                    print(f"Skipping invalid item in JSON: {item}")
                    continue
                
                url = item['url']
                print(f"\n[{i}/{total}]")
                if process_url(url):
                    success_count += 1
            
            print("\n--- Batch Run Complete ---")
            print(f"Successfully processed: {success_count} / {total}")

        except FileNotFoundError:
            print(f"Error: JSON file not found at '{argument}'")
        except json.JSONDecodeError:
            print(f"Error: Could not decode JSON from '{argument}'. Check for syntax errors.")
        except Exception as e:
            print(f"An error occurred during batch processing: {e}")
    
    else:
        # Single URL Mode
        process_url(argument)
